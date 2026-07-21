import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config import REPORTS_DIR
from database import get_all_objects_transactions, get_hq_transactions, get_objects, get_object_transactions, get_object_balance_before, get_object, get_hq_balance_before, get_deleted_operations, EXPENSE_TYPES
from utils import format_date, format_amount, today_str, TZ

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)

TYPE_NAMES = {
    "income": "Приход",
    "expense": "Расход",
    "director_expense": "Расход Директора",
    "supplier_payment": "Оплата Поставщика",
    "transfer_out": "Перевод в офис",
    "transfer_in": "Перевод из объекта",
}


def _style_header(ws, row: int, cols: int):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _amount_cell(ws, row: int, col: int, value: float):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = '#,##0.00'
    cell.border = THIN_BORDER
    return cell


def _auto_width(ws, cols: int, min_width: int = 12, max_width: int = 40):
    for col in range(1, cols + 1):
        max_len = min_width
        for r in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in r:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[get_column_letter(col)].width = max_len + 2


def _add_title(ws, title: str, cols: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30


def _write_transactions(ws, transactions, start_row: int, headers: list[str]):
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=col_idx, value=h)
    _style_header(ws, start_row, len(headers))
    row = start_row + 1

    for t in transactions:
        ws.cell(row=row, column=1, value=t["id"]).border = THIN_BORDER
        dt_str = t["transaction_date"]
        ws.cell(row=row, column=2, value=format_date(dt_str[:10])).border = THIN_BORDER
        ws.cell(row=row, column=3, value=dt_str[11:19] if len(dt_str) > 10 else "").border = THIN_BORDER
        ws.cell(row=row, column=4, value=TYPE_NAMES.get(t["type"], t["type"])).border = THIN_BORDER
        amount_val = -t["amount"] if t["type"] in EXPENSE_TYPES else t["amount"]
        cell = _amount_cell(ws, row, 5, amount_val)
        if t["type"] in EXPENSE_TYPES:
            cell.font = Font(color="FF0000")
        elif t["type"] in ("income", "transfer_in"):
            cell.font = Font(color="008000")
        ws.cell(row=row, column=6, value=t.get("reason", "") or "").border = THIN_BORDER
        ws.cell(row=row, column=7, value=t.get("user_name", "") or "").border = THIN_BORDER
        ws.cell(row=row, column=8, value=t.get("object_name", "") or "").border = THIN_BORDER
        row += 1
    return row


def _write_deleted_operations_sheet(wb, deleted_ops, title: str):
    ws = wb.create_sheet(title="Удаленные")
    del_headers = ["ID", "Дата", "Время", "Тип", "Сумма", "Причина", "Сотрудник", "Объект", "Дата удаления", "Время удаления", "Кто удалил", "Причина удаления"]
    _add_title(ws, title, len(del_headers))

    for col_idx, h in enumerate(del_headers, 1):
        ws.cell(row=2, column=col_idx, value=h)
    _style_header(ws, 2, len(del_headers))

    row = 3
    for d in deleted_ops:
        ws.cell(row=row, column=1, value=d["original_id"]).border = THIN_BORDER
        dt_str = d["transaction_date"]
        ws.cell(row=row, column=2, value=format_date(dt_str[:10])).border = THIN_BORDER
        ws.cell(row=row, column=3, value=dt_str[11:19] if len(dt_str) > 10 else "").border = THIN_BORDER
        ws.cell(row=row, column=4, value=TYPE_NAMES.get(d["type"], d["type"])).border = THIN_BORDER
        amount_val = -d["amount"] if d["type"] in EXPENSE_TYPES else d["amount"]
        cell = _amount_cell(ws, row, 5, amount_val)
        if d["type"] in EXPENSE_TYPES:
            cell.font = Font(color="FF0000")
        elif d["type"] in ("income", "transfer_in"):
            cell.font = Font(color="008000")
        ws.cell(row=row, column=6, value=d.get("reason", "") or "").border = THIN_BORDER
        ws.cell(row=row, column=7, value=d.get("original_user_name", "") or "").border = THIN_BORDER
        ws.cell(row=row, column=8, value=d.get("object_name", "") or "").border = THIN_BORDER
        del_str = d["deleted_at"]
        ws.cell(row=row, column=9, value=format_date(del_str[:10])).border = THIN_BORDER
        ws.cell(row=row, column=10, value=del_str[11:19] if len(del_str) > 10 else "").border = THIN_BORDER
        ws.cell(row=row, column=11, value=d.get("deleted_by_name", "") or "").border = THIN_BORDER
        ws.cell(row=row, column=12, value=d.get("delete_reason", "") or "").border = THIN_BORDER
        row += 1

    _auto_width(ws, len(del_headers))


def _balance_as_of(object_id: int, transactions: list[dict], date_str: str, opening: float) -> float:
    """Calculate balance including all transactions up to and including date_str."""
    bal = opening
    for t in transactions:
        if t["transaction_date"][:10] > date_str:
            break
        if t["type"] == "income":
            bal += t["amount"]
        elif t["type"] in EXPENSE_TYPES:
            bal -= t["amount"]
    return bal


def _extract_metrics(txns: list[dict]) -> dict:
    income = 0.0
    expense = 0.0
    supplier_payment = 0.0
    director_expense = 0.0
    transfer_out = 0.0
    transfer_in = 0.0
    for t in txns:
        type_ = t["type"]
        if type_ == "income":
            income += t["amount"]
        elif type_ == "expense":
            expense += t["amount"]
        elif type_ == "supplier_payment":
            supplier_payment += t["amount"]
        elif type_ == "director_expense":
            director_expense += t["amount"]
        elif type_ == "transfer_out":
            transfer_out += t["amount"]
        elif type_ == "transfer_in":
            transfer_in += t["amount"]
    return {
        "income": income,
        "expense": expense,
        "supplier_payment": supplier_payment,
        "director_expense": director_expense,
        "transfer_out": transfer_out,
        "transfer_in": transfer_in,
    }


# --- General report: one sheet per object ---

async def generate_all_objects_report(start_date: str, end_date: str) -> str:
    os.makedirs(REPORTS_DIR, exist_ok=True)
    filename = f"Общий_отчет_{start_date}_по_{end_date}_{datetime.now(TZ).strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)

    wb = Workbook()
    wb.remove(wb.active)
    objects = await get_objects()
    headers = ["ID", "Дата", "Время", "Тип операции", "Сумма", "Причина", "Сотрудник", "Объект"]

    entity_data = []

    for obj in objects:
        ws = wb.create_sheet(title=obj["name"][:31])
        txns = await get_object_transactions(obj["id"], start_date, end_date)
        opening = await get_object_balance_before(obj["id"], start_date)
        closing = _balance_as_of(obj["id"], txns, end_date, opening)
        metrics = _extract_metrics(txns)

        entity_data.append({
            "name": obj["name"],
            "metrics": metrics,
            "opening": opening,
            "closing": closing,
        })

        title = f"«{obj['name']}» {format_date(start_date)} — {format_date(end_date)}"
        _add_title(ws, title, len(headers))

        ws.cell(row=3, column=1, value="Остаток на начало:").font = Font(bold=True)
        _amount_cell(ws, 3, 2, opening)
        _write_transactions(ws, txns, 5, headers)

        last_row = ws.max_row + 1
        ws.cell(row=last_row, column=1, value="Остаток на конец:").font = Font(bold=True)
        _amount_cell(ws, last_row, 2, closing)

        _auto_width(ws, len(headers))

    # HQ sheet
    hq_headers = ["ID", "Дата", "Время", "Тип операции", "Сумма", "Причина", "Сотрудник", "Источник"]
    ws_hq = wb.create_sheet(title="Головной офис")
    hq_txns = await get_hq_transactions(start_date, end_date)
    hq_opening = await get_hq_balance_before(start_date)
    hq_closing = hq_opening
    for t in hq_txns:
        if t["type"] == "transfer_in":
            hq_closing += t["amount"]
        elif t["type"] in EXPENSE_TYPES:
            hq_closing -= t["amount"]

    hq_metrics = _extract_metrics(hq_txns)

    _add_title(ws_hq, f"Головной офис {format_date(start_date)} — {format_date(end_date)}", len(hq_headers))
    ws_hq.cell(row=3, column=1, value="Остаток на начало:").font = Font(bold=True)
    _amount_cell(ws_hq, 3, 2, hq_opening)

    hq_data_row = 5
    for col_idx, h in enumerate(hq_headers, 1):
        ws_hq.cell(row=hq_data_row, column=col_idx, value=h)
    _style_header(ws_hq, hq_data_row, len(hq_headers))
    r = hq_data_row + 1
    for t in hq_txns:
        ws_hq.cell(row=r, column=1, value=t["id"]).border = THIN_BORDER
        dt_str = t["transaction_date"]
        ws_hq.cell(row=r, column=2, value=format_date(dt_str[:10])).border = THIN_BORDER
        ws_hq.cell(row=r, column=3, value=dt_str[11:19] if len(dt_str) > 10 else "").border = THIN_BORDER
        ws_hq.cell(row=r, column=4, value=TYPE_NAMES.get(t["type"], t["type"])).border = THIN_BORDER
        amount_val = -t["amount"] if t["type"] in EXPENSE_TYPES else t["amount"]
        cell = _amount_cell(ws_hq, r, 5, amount_val)
        if t["type"] in EXPENSE_TYPES:
            cell.font = Font(color="FF0000")
        elif t["type"] == "transfer_in":
            cell.font = Font(color="008000")
        ws_hq.cell(row=r, column=6, value=t.get("reason", "") or "").border = THIN_BORDER
        ws_hq.cell(row=r, column=7, value=t.get("user_name", "") or "").border = THIN_BORDER
        ws_hq.cell(row=r, column=8, value=t.get("source_object_name", "") or "").border = THIN_BORDER
        r += 1

    last_row = ws_hq.max_row + 1
    ws_hq.cell(row=last_row, column=1, value="Остаток на конец:").font = Font(bold=True)
    _amount_cell(ws_hq, last_row, 2, hq_closing)
    _auto_width(ws_hq, len(hq_headers))

    # Summary sheet
    num_cols = 2 + 1 + len(objects)
    ws = wb.create_sheet(title="Сводка")
    _add_title(ws, f"Сводка {format_date(start_date)} — {format_date(end_date)}", num_cols)

    # Header row
    ws.cell(row=3, column=1, value="Показатель")
    ws.cell(row=3, column=2, value="Всего")
    col = 3
    ws.cell(row=3, column=col, value="Головной офис")
    col += 1
    for ed in entity_data:
        ws.cell(row=3, column=col, value=ed["name"])
        col += 1
    _style_header(ws, 3, num_cols)

    BLUE_FONT = Font(color="0000FF", bold=True)
    GREEN_FONT = Font(color="008000")
    RED_FONT = Font(color="FF0000")

    def _write_metric(ws, row, value, is_expense=False, is_balance=False, green=False):
        if is_expense:
            cell = _amount_cell(ws, row, col, -value)
            cell.font = RED_FONT
        elif green:
            cell = _amount_cell(ws, row, col, value)
            cell.font = GREEN_FONT
        else:
            cell = _amount_cell(ws, row, col, value)
            if is_balance:
                cell.font = BLUE_FONT

    indicator_defs = [
        ("Всего приход", "income"),
        ("Всего расход", "expense"),
        ("Всего оплат поставщика", "supplier_payment"),
        ("Всего расход директора", "director_expense"),
        ("Всего переводов в офис", "transfer_out"),
    ]

    r = 4

    # Остаток на начало периода (blue, first row)
    ws.cell(row=r, column=1, value="Остаток на начало периода").border = THIN_BORDER
    ws.cell(row=r, column=1).font = BLUE_FONT
    total_opening = hq_opening + sum(ed["opening"] for ed in entity_data)
    col = 2; _write_metric(ws, r, total_opening, is_balance=True)
    col = 3; _write_metric(ws, r, hq_opening, is_balance=True)
    col = 4
    for ed in entity_data:
        _write_metric(ws, r, ed["opening"], is_balance=True)
        col += 1
    r += 1

    # Indicator rows
    for label, key in indicator_defs:
        is_expense = key != "income"
        ws.cell(row=r, column=1, value=label).border = THIN_BORDER

        if key == "transfer_out":
            hq_val = hq_metrics["transfer_in"]
            objects_val = sum(ed["metrics"]["transfer_out"] for ed in entity_data)
            signed_total = hq_val - objects_val
            col = 2; _write_metric(ws, r, abs(signed_total), is_expense=(signed_total < 0), green=(signed_total > 0))
            col = 3; _write_metric(ws, r, hq_val, is_expense=False, green=True)
            col = 4
            for ed in entity_data:
                _write_metric(ws, r, ed["metrics"]["transfer_out"], is_expense=True)
                col += 1
        else:
            hq_val = hq_metrics[key]
            objects_val = sum(ed["metrics"][key] for ed in entity_data)
            total = hq_val + objects_val
            is_green = key == "income"
            col = 2; _write_metric(ws, r, total, is_expense, green=is_green)
            col = 3; _write_metric(ws, r, hq_val, is_expense, green=is_green)
            col = 4
            for ed in entity_data:
                _write_metric(ws, r, ed["metrics"][key], is_expense, green=is_green)
                col += 1
        r += 1

    # Остаток на конец периода (blue, last row)
    ws.cell(row=r, column=1, value="Остаток на конец периода").border = THIN_BORDER
    ws.cell(row=r, column=1).font = BLUE_FONT
    total_closing = hq_closing + sum(ed["closing"] for ed in entity_data)
    col = 2; _write_metric(ws, r, total_closing, is_balance=True)
    col = 3; _write_metric(ws, r, hq_closing, is_balance=True)
    col = 4
    for ed in entity_data:
        _write_metric(ws, r, ed["closing"], is_balance=True)
        col += 1

    _auto_width(ws, num_cols)

    # Deleted operations sheet
    deleted_ops = await get_deleted_operations(start_date=start_date, end_date=end_date)
    if deleted_ops:
        _write_deleted_operations_sheet(wb, deleted_ops, f"Удаленные операции {format_date(start_date)} — {format_date(end_date)}")

    wb.save(filepath)
    return filepath


# --- Single object report ---

async def generate_object_report(object_id: int, object_name: str, start_date: str, end_date: str) -> str:
    os.makedirs(REPORTS_DIR, exist_ok=True)
    filename = f"Отчет_{object_name}_{start_date}_по_{end_date}_{datetime.now(TZ).strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)

    txns = await get_object_transactions(object_id, start_date, end_date)
    opening = await get_object_balance_before(object_id, start_date)
    closing = _balance_as_of(object_id, txns, end_date, opening)

    wb = Workbook()
    ws = wb.active
    ws.title = object_name[:31]

    headers = ["ID", "Дата", "Время", "Тип операции", "Сумма", "Причина", "Сотрудник", "Объект"]
    _add_title(ws, f"«{object_name}» {format_date(start_date)} — {format_date(end_date)}", len(headers))

    ws.cell(row=3, column=1, value="Остаток на начало периода:").font = Font(bold=True)
    _amount_cell(ws, 3, 2, opening)

    _write_transactions(ws, txns, 5, headers)

    last_row = ws.max_row + 1
    ws.cell(row=last_row, column=1, value="Остаток на конец периода:").font = Font(bold=True)
    _amount_cell(ws, last_row, 2, closing)

    _auto_width(ws, len(headers))

    # Deleted operations sheet
    del_ops = await get_deleted_operations(object_id=object_id, start_date=start_date, end_date=end_date)
    if del_ops:
        _write_deleted_operations_sheet(wb, del_ops, f"Удаленные операции «{object_name}» {format_date(start_date)} — {format_date(end_date)}")

    wb.save(filepath)
    return filepath


# --- Daily report ---

async def generate_daily_report(object_id: int, object_name: str, date_str: str | None = None) -> str:
    if date_str is None:
        date_str = today_str()
    os.makedirs(REPORTS_DIR, exist_ok=True)
    filename = f"Дневной_{object_name}_{date_str}_{datetime.now(TZ).strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)

    from database import get_daily_summary

    summary = await get_daily_summary(object_id, date_str)
    txns = await get_object_transactions(object_id, date_str, date_str)
    headers = ["ID", "Дата", "Время", "Тип операции", "Сумма", "Причина", "Сотрудник", "Объект"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Дневной отчет"
    _add_title(ws, f"Дневной отчет «{object_name}» за {format_date(date_str)}", len(headers))

    ws.cell(row=3, column=1, value="Остаток на начало дня:").font = Font(bold=True)
    _amount_cell(ws, 3, 2, summary["opening"])

    _write_transactions(ws, txns, 5, headers)

    r = ws.max_row + 1
    ws.cell(row=r, column=1, value="Итого приход:").font = Font(bold=True)
    _amount_cell(ws, r, 2, summary["income"])
    r += 1
    ws.cell(row=r, column=1, value="Итого расход:").font = Font(bold=True)
    _amount_cell(ws, r, 2, summary["expense"])
    r += 1
    ws.cell(row=r, column=1, value="Перевод в офис:").font = Font(bold=True)
    _amount_cell(ws, r, 2, summary["transfer_out"])
    r += 1
    ws.cell(row=r, column=1, value="Остаток на конец дня:").font = Font(bold=True)
    _amount_cell(ws, r, 2, summary["closing"])

    _auto_width(ws, len(headers))

    # Deleted operations sheet
    del_ops = await get_deleted_operations(object_id=object_id, start_date=date_str, end_date=date_str)
    if del_ops:
        _write_deleted_operations_sheet(wb, del_ops, f"Удаленные операции «{object_name}» за {format_date(date_str)}")

    wb.save(filepath)
    return filepath


# --- Text summary for Telegram ---

async def generate_object_report_text(object_id: int, object_name: str, start_date: str, end_date: str) -> str:
    txns = await get_object_transactions(object_id, start_date, end_date)
    opening = await get_object_balance_before(object_id, start_date)
    closing = _balance_as_of(object_id, txns, end_date, opening)
    m = _extract_metrics(txns)

    lines = [f"📅 {format_date(start_date)} — {format_date(end_date)}"]
    lines.append(f"🔵 Остаток на начало: {format_amount(opening)} сум")
    lines.append(f"🟢 Всего приход: +{format_amount(m['income'])} сум")
    lines.append(f"🔴 Всего расход: -{format_amount(m['expense'])} сум")
    lines.append(f"📦 Оплата поставщика: -{format_amount(m['supplier_payment'])} сум")
    lines.append(f"👔 Расход директора: -{format_amount(m['director_expense'])} сум")
    lines.append(f"💸 Перевод в офис: -{format_amount(m['transfer_out'])} сум")
    lines.append(f"🏁 Остаток на конец: {format_amount(closing)} сум")
    return "\n".join(lines)
